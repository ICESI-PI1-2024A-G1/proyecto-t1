from datetime import datetime
from io import BytesIO
import os
from openpyxl import Workbook
import pandas as pd
from django.http import Http404, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook
from django.conf import settings
from openpyxl.styles import Alignment
import xlwings as xw
import tempfile


class SharePointAPI:
    """Class to control the information comming from the sharepoint api"""

    def __init__(self, excel_path) -> None:
        """
        Initializes the SharePointAPI object.

        Args:
            excel_path (str): The path to the Excel file.
        """
        self.excel_path = excel_path
        self.request_columns = [
            "id",
            "status",
            "manager",
            "team",
            "initial_date",
            "final_date",
            "fullname",
            "faculty",
            "document",
            "phone_number",
            "email",
            "CENCO",
            "bank",
            "account_type",
            "health_provider",
            "pension_fund",
            "arl",
            "contract_value",
            "is_one_time_payment",
        ]

    def clear_db(self, showLog = True):
        """
        Clears the database by removing the Excel file and creating a new one.
        """
        try:
            os.remove(self.excel_path)
            if showLog:
                print("Database clear")
        except:
            pass
        if not os.path.exists(os.path.dirname(self.excel_path)):
            try:
                os.makedirs(os.path.dirname(self.excel_path))
                print("Directories created")
            except Exception as e:
                print(f"Error creating directories: {e}")
                return
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "data"
        workbook.save(self.excel_path)
        workbook.close()

    @csrf_exempt
    def get_request_by_id(self, id) -> JsonResponse:
        """
        Retrieves a request by its ID from the Excel data.

        Args:
            id (int): The ID of the request.

        Returns:
            JsonResponse: JSON response containing the request data.

        Raises:
            Http404: If the request is not found.
        """
        try:
            df = pd.read_excel(self.excel_path, sheet_name="data", header=0)
            data = df[df["id"] == id].to_dict(orient="records")
            if len(data) > 0:
                return JsonResponse(data=data[0], status=200, safe=False)
            else:
                raise Http404("Solicitud no encontrada.")
        except FileNotFoundError:
            raise Http404("El archivo no se encontró")

    @csrf_exempt
    def get_all_requests(self) -> JsonResponse:
        """
        Retrieves all requests from the Excel data.

        Returns:
            JsonResponse: JSON response containing all request data.

        Raises:
            Http404: If the file is not found.
        """
        try:
            df = pd.read_excel(self.excel_path, sheet_name="data", header=0)
            data = df.to_dict(orient="records")
            return JsonResponse(data, status=200, safe=False)
        except FileNotFoundError:
            raise Http404("El archivo no se encontró")

    @csrf_exempt
    def create_data(self, data) -> JsonResponse:
        """
        Creates new data in the Excel file.

        Args:
            data (dict): The data to be created.

        Returns:
            JsonResponse: JSON response confirming the creation.

        Raises:
            Http404: If the request creation fails.
        """
        try:
            excel_file = pd.ExcelFile(self.excel_path)
            df = pd.read_excel(self.excel_path, sheet_name="data")
            start_row = len(df)
            new_data = data.copy()
            new_data["id"] = start_row + 1
            new_df = pd.DataFrame([new_data])
            result = pd.concat([df, new_df], ignore_index=True)
            result.to_excel(
                self.excel_path,
                sheet_name="data",
                index=False,
                columns=self.request_columns,
            )
            return JsonResponse(
                {"mensaje": "Información creada correctamente"}, status=201, safe=False
            )
        except Exception as e:
            raise Http404("No se pudo crear la solicitud.")

    @csrf_exempt
    def update_data(self, id, new_data) -> JsonResponse:
        """
        Updates existing data in the Excel file.

        Args:
            id (int): The ID of the data to be updated.
            new_data (dict): The updated data.

        Returns:
            JsonResponse: JSON response confirming the update.

        Raises:
            Http404: If the request update fails.
        """
        try:
            df = pd.read_excel(self.excel_path, sheet_name="data", header=0)
            mask = df["id"] == id
            if not df[mask].empty:
                for column_name, new_value in new_data.items():
                    df.loc[mask, column_name] = new_value
                df.to_excel(self.excel_path, sheet_name="data", index=False)
                return JsonResponse(
                    {"mensaje": "Dato actualizado satisfactoriamente"},
                    status=200,
                    safe=False,
                )
            else:
                raise Http404("Solicitud no encontrada.")
        except FileNotFoundError:
            raise Http404("El archivo no se encontró")

    @csrf_exempt
    def delete_data(self, id) -> JsonResponse:
        """
        Deletes data from the Excel file.

        Args:
            id (int): The ID of the data to be deleted.

        Returns:
            JsonResponse: JSON response confirming the deletion.

        Raises:
            Http404: If the data deletion fails.
        """
        try:
            df = pd.read_excel(self.excel_path, sheet_name="data", header=0)
            original_rows = len(df)
            df = df[df["id"] != id]
            new_rows = len(df)
            if new_rows < original_rows:
                df.to_excel(self.excel_path, sheet_name="data", index=False)
                return JsonResponse(
                    {"mensaje": "Dato eliminado satisfactoriamente"},
                    status=200,
                    safe=False,
                )
            else:
                raise Http404(
                    "No se encontró la información asociada al ID proporcionado"
                )
        except FileNotFoundError:
            raise Http404("El archivo no se encontró")

    @csrf_exempt
    def search_data(self, query) -> JsonResponse:
        """
        Searches data in the Excel file based on a query.

        Args:
            query (str or dict): The query string or dictionary.

        Returns:
            JsonResponse: JSON response containing the search results.

        Raises:
            Http404: If the search fails.
        """
        try:
            df = pd.read_excel(self.excel_path, sheet_name="data", header=0)
            filtered_df = df.copy()
            if query == "None":
                return self.get_all_requests()
            if isinstance(query, str):
                query = query.lower()
                filtered_df = filtered_df[
                    filtered_df.apply(
                        lambda x: x.astype(str).str.lower().str.contains(query), axis=1
                    ).any(axis=1)
                ]
            else:
                filtered_df = filtered_df[
                    filtered_df.apply(lambda x: x == query, axis=1).any(axis=1)
                ]
            ans = filtered_df.to_dict(orient="records")
            return JsonResponse(ans, status=200, safe=False)
        except FileNotFoundError:
            raise Http404("El archivo no se encontró")

    @csrf_exempt
    def remove_team(self, id) -> JsonResponse:
        """
        Removes a team from the data by setting its team ID to NaN.

        Args:
            id (int): The ID of the team to be removed.

        Returns:
            JsonResponse: JSON response confirming the removal.

        Raises:
            Exception: If the team removal fails.
        """
        try:
            # Cargar el archivo Excel
            df = pd.read_excel(self.excel_path, sheet_name="data", header=0)

            # Reemplazar el ID del equipo con NaN en el campo "team" del Excel
            df.loc[df["team"] == id, "team"] = pd.NA

            # Guardar el archivo Excel actualizado
            df.to_excel(
                self.excel_path,
                sheet_name="data",
                index=False,
                columns=self.request_columns,
            )

            return JsonResponse(
                {"mensaje": f"Equipo con ID {id} removido correctamente"},
                status=200,
                safe=False,
            )
        except Exception as e:
            return JsonResponse(
                {"error": str(e)},
                status=500,
                safe=False,
            )

    # def get_form_render(self, form_name=None, excel_file=None):
    #     try:
    #         if excel_file:
    #             content = excel_file.read()
    #             wb = load_workbook(filename=BytesIO(content))
    #         else:
    #             excel_file_path = os.path.join(self.excel_path, form_name)
    #             wb = load_workbook(excel_file_path)
    #         sheet = wb.active
    #         data = []

    #         # Obtener información sobre celdas combinadas
    #         merged_cells_ranges = sheet.merged_cells.ranges
    #         merged_cells_map = {}  # Mapa para rastrear las celdas combinadas
    #         spans_map = {}

    #         for merged_range in merged_cells_ranges:
    #             min_row, min_col, max_row, max_col = (
    #                 merged_range.min_row,
    #                 merged_range.min_col,
    #                 merged_range.max_row,
    #                 merged_range.max_col,
    #             )
    #             spans_map[(min_row, min_col)] = (
    #                 max_row - min_row + 1,
    #                 max_col - min_col + 1,
    #             )
    #             for row in range(min_row, max_row + 1):
    #                 for col in range(min_col, max_col + 1):
    #                     merged_cells_map[(row, col)] = (min_row, min_col)

    #         for row in sheet.iter_rows(values_only=False):
    #             row_data = []
    #             for cell in row:
    #                 row_idx, col_idx = cell.row, cell.column
    #                 display = True
    #                 row_span = col_span = 1
    #                 if (row_idx, col_idx) in merged_cells_map:
    #                     # Si la celda está combinada, usar el valor de la celda de origen
    #                     origin_row, origin_col = merged_cells_map[(row_idx, col_idx)]
    #                     is_origin = row_idx == origin_row and col_idx == origin_col
    #                     row_idx = origin_row
    #                     col_idx = origin_col
    #                     if not is_origin:
    #                         display = False
    #                 if (row_idx, col_idx) in spans_map:
    #                     row_span, col_span = spans_map[(row_idx, col_idx)]
    #                 row_data.append(
    #                     {
    #                         "value": cell.value if cell.value is not None else "",
    #                         "row_span": row_span,
    #                         "col_span": col_span,
    #                         "row_idx": row_idx,
    #                         "col_idx": col_idx,
    #                         "display": display,
    #                     }
    #                 )
    #             data.append(row_data)
    #         return JsonResponse(data, status=200, safe=False)
    #     except FileNotFoundError:
    #         raise Http404("El archivo no se encontró")

    # def fill_form(self, excel_file, form_fields):
    #     wb = xw.Book(excel_file.path)
    #     wb.app.visible = False
    #     sheet = wb.sheets[0]

    #     for field in form_fields:
    #         row_idx = int(field["row_idx"])
    #         col_idx = int(field["col_idx"])
    #         cell_value = field["value"]
    #         cell = sheet.cells(row_idx, col_idx)
    #         cell.value = cell_value
    #         cell.api.HorizontalAlignment = -4108
    #         cell.api.VerticalAlignment = -4108

    #     curr_date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    #     output_path = os.path.join(settings.MEDIA_ROOT, "filled_forms")
    #     if not os.path.exists(output_path):
    #         os.makedirs(output_path)
    #     output_file_path = os.path.join(output_path, f"filled-{curr_date}.xlsx")
    #     wb.save(output_file_path)
    #     wb.close()

    #     return output_file_path